import tkinter as tk
from tkinter import ttk
import requests
from bs4 import BeautifulSoup
import lxml.html
import openpyxl
import shutil
import os, tkinter, tkinter.filedialog, tkinter.messagebox
import threading

def download_img(url, file_name):
    r = requests.get(url, stream=True)
    if r.status_code == 200:
        with open(file_name, 'wb') as f:
            r.raw.decode_content = True
            shutil.copyfileobj(r.raw, f)

def function():
    #Excel開く
    root = tkinter.Tk()
    root.withdraw()
    fTyp = [("", "*")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    # tkinter.messagebox.showinfo('○×プログラム','処理ファイルを選択してください！')
    file = tkinter.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
    #file = 'input.xlsx'
    wb1 = openpyxl.load_workbook(file)
    ws1 = wb1.worksheets[0]
    values = []
    for cell in ws1['A']:
        values.append(cell.value)
    #print(values)

    n = 2
    book = openpyxl.Workbook()
    sheet = book['Sheet']

    title = ["メーカー名", "ブランド名", "商品名", "商品説明文", "成分特徴", "使い方", "使用上の注意", "商品公式URL", "JANコード", "色番号・色名", "バリエーション説明文",
             "価格・容量・単位", "SPF/PA", "全成分"]
    for i in range(0, 14):
        sheet.cell(1, i + 1).value = title[i]
        # sheet.cell(1, i+1).alignment = openpyxl.styles.Alignment(wrapText=True)
    for i in range(14,24):
        sheet.cell(1, i + 1).value = '画像URL' + str(i-13)
    for i in range(24,34):
        sheet.cell(1, i + 1).value = '画像ファイル名' + str(i-23)
    sheet.column_dimensions['A'].width = '15'
    sheet.column_dimensions['B'].width = '15'
    sheet.column_dimensions['C'].width = '15'
    sheet.column_dimensions['D'].width = '15'
    sheet.column_dimensions['E'].width = '15'
    sheet.column_dimensions['F'].width = '15'
    sheet.column_dimensions['G'].width = '15'
    sheet.column_dimensions['H'].width = '15'
    sheet.column_dimensions['I'].width = '15'
    sheet.column_dimensions['J'].width = '15'
    sheet.column_dimensions['K'].width = '15'
    sheet.column_dimensions['L'].width = '15'
    sheet.column_dimensions['M'].width = '15'
    sheet.column_dimensions['N'].width = '15'


    #while True:
    for url in values:
        #url = 'https://www.cosme.net/product/product_id/10159831/sku/1003546'
        #url = 'https://www.cosme.net/product/product_id/10144703/top'

        res = requests.get(url)
        soup = BeautifulSoup(res.content, "lxml")

        maker_element = soup.select_one('#product-spec > dl.maker.clearfix > dd > a')
        try:
            maker = maker_element.get_text()
            sheet.cell(n, 1).value = maker
        except:
            print("none")

        bland_element = soup.select_one('#product-spec > dl.brand-name.clearfix > dd > a')
        try:
            bland = bland_element.get_text()
            sheet.cell(n, 2).value = bland
        except:
            print("none")

        """
        goods_element = soup.select_one('#product-header > h2 > strong > a')
        try:
            goods = goods_element.get_text()
            sheet.cell(n, 3).value = goods
        except:
            goods_element = soup.select_one('#product-header > h2 > strong > span > a')
            goods = goods_element.get_text()
            sheet.cell(n, 3).value = goods
            print("none")
        """
        if 'top' in url:
            goods_element = soup.select_one('#product-header > h2 > strong > a')
            goods = goods_element.get_text()
            sheet.cell(n, 3).value = goods
        else:
            goods_element = soup.select_one('#product-header > h2 > strong > span > a')
            goods = goods_element.get_text()
            goods_element2 = soup.select_one('#product-header > h2 > strong > span > span')
            goods2 = goods_element2.get_text()
            goods = goods + goods2
            sheet.cell(n, 3).value = goods

        goods_explanation_element = soup.select_one('#product-spec > dl.item-description.clearfix > dd')
        try:
            goods_explanation = goods_explanation_element.get_text()
            sheet.cell(n, 4).value = goods_explanation
        except:
            print("none")
        #print(goods_explanation)

        component_element = soup.select_one('#product-spec > dl.ingredient.clearfix > dd > ul')
        try:
            component = component_element.get_text()
            sheet.cell(n, 5).value = component
        except:
            print("none")

        usage_element = soup.select_one('#product-spec > dl.use.clearfix > dd')
        try:
            usage = usage_element.get_text()
            sheet.cell(n, 6).value = usage
        except:
            print("none")

        caution_element = soup.select_one('#product-spec > dl.precautions.clearfix > dd')
        try:
            caution = caution_element.get_text()
            #caution = '\n'.join(caution.splitlines()[1:])
            #caution = '\n'.join(caution.splitlines()[1:])
            lines = caution.splitlines()
            if lines[0] == '':
                caution = '\n'.join(caution.splitlines()[1:])
            sheet.cell(n, 7).value = caution
        #print(caution)
        except:
            print("none")

        url2_element = soup.select_one('#product-spec > dl.official-site.clearfix > dd > a')
        try:
            url2 = url2_element.get_attribute_list('href')
            url2 = url2[0]
            sheet.cell(n, 8).value = url2
        except:
            print("none")
        #print(url2)
        if 'top' in url:
            JAN_elements = soup.select('#product-spec > div > dl > dd')
            try:
                JAN = ''
                for JAN_element in JAN_elements:
                    JAN = JAN + JAN_element.get_text() + '\n'
                sheet.cell(n, 9).value = JAN
            except:
                print("none")
        else:
            JAN_elements = soup.select('#product-spec > dl.jan-code.clearfix > dd > ul > li')
            try:
                JAN = ''
                for JAN_element in JAN_elements:
                    JAN = JAN + JAN_element.get_text() + '\n'
                sheet.cell(n, 9).value = JAN
            except:
                print("none")

        color_element = soup.select_one('#product-spec > dl.color.clearfix > dd')
        try:
            color = color_element.get_text()
            sheet.cell(n, 10).value = color
        except:
            print("none")

        price_element = soup.select_one('#product-spec > dl.capacity-and-price.clearfix > dd')
        try:
            price = price_element.get_text()
            sheet.cell(n, 11).value = price
        except:
            print("none")

        SPF_element = soup.select_one('#product-spec > dl.spf.clearfix > dd')
        try:
            SPF = SPF_element.get_text()
            sheet.cell(n, 12).value = SPF
        except:
            print("none")

        all_component_element = soup.select_one('#product-spec > dl.all-components.clearfix > dd')
        try:
            all_component = all_component_element.get_text()
            sheet.cell(n, 13).value = all_component
            #sheet.cell(n, 14).alignment = openpyxl.styles.Alignment(wrapText=True)
        except:
            print("none")

        i = 1
        try:
            while True:
                if 'top' in url:
                    img_element = soup.select_one('#main > div.vri-item > div.vri-item-inr-top > ul > li:nth-child(' + str(i) + ') > a > p.vari-pic > img')
                else:
                    img_element = soup.select_one('#thumb-newdb-1606 > div > div > div > div > ul > li:nth-child(' + str(i) + ') > img')
                img_url = img_element.get('src')
                img_url = img_url.replace('?target=70x70', '')
                img_url = img_url.replace('.jpg', '')
                img_url = img_url + '_xl.jpg?target=350x350&size=trimIfLarge'
                sheet.cell(n, i + 14).value = img_url
                sheet.cell(n, i + 24).value = str(n-1) + '_' + str(i)
                download_img(img_url,'image//' + str(n-1) + '_' + str(i) + '.jpg')
                i = i + 1
        except:
            print('終了')
        save_name.set(str(n-1) + '件目処理完了')
        n = n + 1
    save_name.set('全件処理完了')
    book.save('output.xlsx')
#function()
def clicked():
    thread1 = threading.Thread(target=function)
    thread1.start()

root = tk.Tk()
root.title('@コスメ')
frame2 = ttk.Frame(root, padding=16)
button1 = ttk.Button(
    frame2,
    text='ファイルを選択',
    command=clicked)
frame4= ttk.Frame(root, padding=16)
save_name = tk.StringVar()
entry4 = ttk.Entry(frame4, textvariable=save_name, width = 30)
frame2.pack(side=tk.TOP, anchor=tk.NW)
button1.pack(fill=tk.X)
frame4.pack(side=tk.TOP, anchor=tk.NW)
entry4.pack(side=tk.TOP, anchor=tk.NW)

# ウィンドウの表示開始
root.mainloop()

#長文だと最初に空白の行が挿入されてしまうかも